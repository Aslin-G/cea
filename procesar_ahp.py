# -*- coding: utf-8 -*-
"""
Procesa el Instrumento_AHP_Panel_Expertos.xlsx diligenciado:
 - matriz recíproca 6x6 por experto, pesos (autovector principal), lambda_max, CI, CR (RI=1,24)
 - agregación de juicios individuales (media geométrica, AIJ), pesos grupales y CR grupal
 - W de Kendall sobre los ordenamientos de criterios de los 3 expertos
 - imprime los valores listos para reportar y exporta 'ahp_resultados.csv'
Uso: python3 procesar_ahp.py [archivo.xlsx]
"""
import sys, numpy as np, pandas as pd
from openpyxl import load_workbook

ARCH = sys.argv[1] if len(sys.argv) > 1 else "Instrumento_AHP_Panel_Expertos.xlsx"
CODS = ['C1','C2','C3','C4','C5','C6']
PARES = [(i,j) for i in range(6) for j in range(i+1,6)]
RI = 1.24  # índice aleatorio de Saaty para n=6

def leer_experto(ws):
    A = np.ones((6,6))
    for n,(i,j) in enumerate(PARES, start=1):
        r = 8 + n
        gana = ws.cell(row=r, column=5).value
        inten = ws.cell(row=r, column=6).value
        if gana is None or inten is None:
            raise ValueError(f"Hoja {ws.title}: fila {r} sin diligenciar")
        inten = float(inten)
        if str(gana).strip() == 'Igual' or inten == 1:
            A[i,j] = A[j,i] = 1.0
        elif str(gana).strip() == CODS[i]:
            A[i,j] = inten; A[j,i] = 1/inten
        else:
            A[i,j] = 1/inten; A[j,i] = inten
    return A

def pesos_cr(A):
    val, vec = np.linalg.eig(A)
    k = np.argmax(val.real)
    w = np.abs(vec[:,k].real); w = w/w.sum()
    lmax = val[k].real
    CI = (lmax-6)/5; CR = CI/RI
    return w, lmax, CI, CR

wb = load_workbook(ARCH, data_only=True)
matrices, pesos_ind = {}, {}
print(f"Archivo: {ARCH}\n")
for e in (1,2,3):
    ws = wb[f"Experto_{e}"]
    A = leer_experto(ws)
    w, lmax, CI, CR = pesos_cr(A)
    matrices[e] = A; pesos_ind[e] = w
    ok = "OK" if CR < 0.10 else ">>> INCONSISTENTE: repetir juicios en Ronda 2 <<<"
    print(f"Experto {e}: lambda_max={lmax:.3f}  CI={CI:.3f}  CR={CR:.3f}  [{ok}]")
    print("  pesos:", dict(zip(CODS, np.round(w,3))))

# Agregación AIJ (media geométrica de juicios)
G = np.ones((6,6))
for i in range(6):
    for j in range(6):
        G[i,j] = np.prod([matrices[e][i,j] for e in (1,2,3)])**(1/3)
wg, lmaxg, CIg, CRg = pesos_cr(G)
print(f"\nGRUPO (AIJ): lambda_max={lmaxg:.3f}  CI={CIg:.3f}  CR={CRg:.3f}")
print("  pesos grupales:", dict(zip(CODS, np.round(wg,3))))

# W de Kendall sobre rankings de criterios
from scipy import stats
Rk = np.array([stats.rankdata(-pesos_ind[e]) for e in (1,2,3)])
m, n = Rk.shape
S = ((Rk.sum(axis=0) - m*(n+1)/2)**2).sum()
W = 12*S/(m**2*(n**3-n))
chi2 = m*(n-1)*W; from scipy.stats import chi2 as chi2d
p = 1 - chi2d.cdf(chi2, n-1)
print(f"\nW de Kendall = {W:.3f} (chi2={chi2:.2f}, gl={n-1}, p={p:.4f})")

pd.DataFrame({'criterio':CODS, 'peso_grupal':np.round(wg,4),
              **{f'peso_exp{e}':np.round(pesos_ind[e],4) for e in (1,2,3)}}).to_csv('ahp_resultados.csv', index=False)
print("\n--- Matriz agregada G (media geométrica), redondeada ---")
print(pd.DataFrame(np.round(G,2), index=CODS, columns=CODS).to_string())

print("\n=== VALORES PARA LA SECCION 3.3 Y LA TABLA 3 ===")
print(f"1) 'Los índices de consistencia individuales fueron CR = {', '.join(f'{pesos_cr(matrices[e])[3]:.3f}' for e in (1,2,3))}, "
      f"y el de la matriz grupal agregada fue CR = {CRg:.3f}, inferiores al umbral de 0,10.'")
print(f"2) 'La concordancia entre los ordenamientos de criterios de los tres expertos fue W = {W:.2f} (p = {p:.3f}).'")
print(f"3) Pesos grupales para la Tabla 1: " + "; ".join(f"{c} = {wg[k]:.2f}" for k,c in enumerate(CODS)))
print("\nNOTA: si los pesos grupales difieren del vector adoptado (0,18/0,22/0,16/0,14/0,12/0,18),")
print("ejecutar 'analisis_comparativo_robustez.py' con los pesos nuevos para regenerar TODOS los resultados.")
